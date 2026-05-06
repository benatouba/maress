"""Backfill missing item DOIs from Marcos publications data.

Matches existing DB items by (title, year) and sets DOI when missing.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
from sqlmodel import select

from app import crud
from app.core.db import SessionLocal
from app.models import Item


NA_LIKE_VALUES = {"na", "n/a", "none", "null", "-"}


def _normalize_title(value: str | None) -> str:
    if not value:
        return ""
    return " ".join(value.strip().casefold().split())


def _normalize_year(value: str | None) -> str:
    if not value:
        return ""
    return value.strip()


def _normalize_doi(value: object) -> str | None:
    if value is None:
        return None
    doi = str(value).strip()
    if not doi:
        return None
    if doi.casefold() in NA_LIKE_VALUES:
        return None
    return doi[:128]


def read_doi_map(publications_path: Path) -> tuple[dict[tuple[str, str], str], int]:
    wb = openpyxl.load_workbook(publications_path, read_only=True)
    ws = wb.active
    if ws is None:
        msg = f"No active sheet in {publications_path}"
        raise ValueError(msg)

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    columns = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}

    required = {"title", "year", "doi"}
    missing = required - set(columns.keys())
    if missing:
        msg = f"Missing columns in publications.xlsx: {missing}"
        raise ValueError(msg)

    doi_map: dict[tuple[str, str], str] = {}
    conflicts = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        title = _normalize_title(row[columns["title"]])
        year_raw = row[columns["year"]]
        year = _normalize_year(str(int(year_raw)) if year_raw else None)
        doi = _normalize_doi(row[columns["doi"]])
        if not title or not year or not doi:
            continue

        key = (title, year)
        existing = doi_map.get(key)
        if existing and existing != doi:
            conflicts += 1
            continue
        doi_map[key] = doi

    wb.close()
    return doi_map, conflicts


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Backfill missing DOIs by matching existing rows to Marcos publications (title+year).",
    )
    parser.add_argument(
        "--email",
        required=True,
        help="Email of the user whose items should be updated.",
    )
    parser.add_argument(
        "--data-dir",
        default="../marcos-custom-data",
        help="Path to marcos-custom-data directory (default: ../marcos-custom-data).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Report updates without writing to the database.",
    )
    args = parser.parse_args()

    data_dir = Path(args.data_dir)
    publications_path = data_dir / "publications.xlsx"
    if not publications_path.exists():
        print(f"ERROR: File not found: {publications_path}")
        sys.exit(1)

    doi_map, conflicts = read_doi_map(publications_path)
    if not doi_map:
        print("No DOI mappings found in publications.xlsx")
        return

    print(f"Loaded {len(doi_map)} title+year -> DOI mappings")
    if conflicts:
        print(f"Skipped {conflicts} conflicting DOI rows in publications.xlsx")

    with SessionLocal() as session:
        user = crud.get_user_by_email(session=session, email=args.email)
        if not user:
            print(f"ERROR: No user found with email '{args.email}'")
            sys.exit(1)

        statement = select(Item).where(Item.owner_id == user.id)
        items = session.exec(statement).all()

        scanned = 0
        missing_doi = 0
        matched = 0
        updated = 0

        for item in items:
            scanned += 1
            if item.doi is not None and item.doi.strip() != "":
                continue

            missing_doi += 1
            key = (_normalize_title(item.title), _normalize_year(item.date))
            if not key[0] or not key[1]:
                continue

            doi = doi_map.get(key)
            if not doi:
                continue

            matched += 1
            if not args.dry_run:
                item.doi = doi
                session.add(item)
                updated += 1

        if args.dry_run:
            print("\n--- DRY RUN ---")
        else:
            session.commit()

        print(f"Scanned items:          {scanned}")
        print(f"Items missing DOI:      {missing_doi}")
        print(f"Matched title+year:     {matched}")
        print(f"DOIs updated:           {updated if not args.dry_run else matched}")


if __name__ == "__main__":
    main()
