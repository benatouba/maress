"""Import Marcos' curated publication and study site data into the Maress
database."""

from __future__ import annotations

import argparse
import logging
import random
import re
import string
import sys
import time
from pathlib import Path
from urllib.parse import quote

import geopandas as gpd
import httpx
import openpyxl
from sqlmodel import select

from app import crud
from app.core.db import SessionLocal
from app.models import CreatorCreate, Item, ItemCreate, StudySiteCreate
from maress_types import CoordinateExtractionMethod, CoordinateSourceType, PaperSections

logger = logging.getLogger(__name__)

CROSSREF_API_BASE = "https://api.crossref.org/works"
CROSSREF_POLITE_DELAY = 0.5  # seconds between requests (polite pool)


def generate_key() -> str:
    """Generate a random 8-character [A-Z0-9] key."""
    return "".join(random.choices(string.ascii_uppercase + string.digits, k=8))  # noqa: S311


def parse_authors(authors_year: str) -> list[dict[str, str]]:
    """Parse Authors_year string like 'Araya_Lopez_et_al2018' into author
    records.

    Returns list of dicts with 'lastName' and 'creatorType' keys.
    """
    # Remove trailing year (4 digits at end)
    name_part = re.sub(r"\d{4}$", "", authors_year).strip("_")
    if not name_part:
        return []

    # Split on _et_al (remaining part is just "et al", skip it)
    has_et_al = "_et_al" in name_part
    name_part = re.split(r"_et_al", name_part)[0]

    # Split on _and_ for multiple authors
    author_parts = re.split(r"_and_", name_part)

    authors = []
    for part in author_parts:
        # Replace remaining underscores with spaces for multi-word last names
        last_name = part.replace("_", " ").strip()
        if last_name:
            authors.append({"lastName": last_name, "creatorType": "author"})

    # If there was "et al", add a placeholder
    if has_et_al:
        authors.append({"lastName": "et al.", "creatorType": "author"})

    return authors


def _crossref_headers(email: str) -> dict[str, str]:
    """Build polite CrossRef API headers."""
    return {
        "User-Agent": f"Maress/0.1 (mailto:{email})",
        "Accept": "application/json",
    }


def _parse_crossref_work(work: dict) -> dict:
    """Extract relevant fields from a CrossRef work object."""
    title = None
    if work.get("title"):
        title = work["title"][0][:255]

    abstract = work.get("abstract", "") or ""
    # CrossRef abstracts may contain JATS XML tags — strip them
    abstract = re.sub(r"<[^>]+>", "", abstract)[:8192]

    doi = work.get("DOI", "")[:128] or None

    year = None
    for date_field in ("published-print", "published-online", "issued"):
        date_parts = work.get(date_field, {}).get("date-parts", [[]])
        if date_parts and date_parts[0] and date_parts[0][0]:
            year = str(date_parts[0][0])
            break

    authors = []
    for author in work.get("author", []):
        family = author.get("family", "")
        given = author.get("given")
        if family:
            authors.append({
                "lastName": family,
                "firstName": given,
                "creatorType": "author",
            })

    return {
        "title": title,
        "doi": doi,
        "abstract": abstract,
        "year": year,
        "authors": authors,
    }


def fetch_crossref_by_doi(doi: str, *, email: str) -> dict | None:
    """Fetch metadata from CrossRef by DOI.

    Args:
        doi: The DOI string (e.g. "10.1234/example").
        email: Contact email for polite pool.

    Returns:
        Parsed metadata dict, or None on failure.
    """
    url = f"{CROSSREF_API_BASE}/{quote(doi, safe='')}"
    try:
        resp = httpx.get(url, headers=_crossref_headers(email), timeout=15)
        if resp.status_code == 200:  # noqa: PLR2004
            work = resp.json().get("message", {})
            return _parse_crossref_work(work)
        logger.warning("CrossRef lookup by DOI %s returned %d", doi, resp.status_code)
    except httpx.HTTPError:
        logger.warning("CrossRef request failed for DOI %s", doi, exc_info=True)
    return None


def fetch_crossref_by_title(title: str, year: str | None, *, email: str) -> dict | None:
    """Search CrossRef by title (and optionally year) and return best match.

    Args:
        title: Paper title to search.
        year: Publication year for filtering (optional).
        email: Contact email for polite pool.

    Returns:
        Parsed metadata dict, or None if no good match.
    """
    params: dict[str, str | int] = {
        "query.bibliographic": title,
        "rows": 3,
    }
    if year:
        params["filter"] = f"from-pub-date:{year},until-pub-date:{year}"

    try:
        resp = httpx.get(
            CROSSREF_API_BASE,
            params=params,
            headers=_crossref_headers(email),
            timeout=15,
        )
        if resp.status_code != 200:  # noqa: PLR2004
            logger.warning("CrossRef title search returned %d", resp.status_code)
            return None

        items = resp.json().get("message", {}).get("items", [])
        if not items:
            return None

        # Take the first (highest relevance) result
        return _parse_crossref_work(items[0])
    except httpx.HTTPError:
        logger.warning("CrossRef title search failed for '%s'", title, exc_info=True)
    return None


def enrich_publication(pub: dict, *, email: str) -> tuple[dict, bool]:
    """Fill missing publication fields using CrossRef.

    Tries DOI lookup first, then title+year search as fallback.

    Args:
        pub: Publication dict with keys: title, doi, abstract, year, authors_year.
        email: Contact email for CrossRef polite pool.

    Returns:
        Tuple of (enriched pub dict, whether any field was updated).
    """
    needs_title = not pub.get("title")
    needs_abstract = not pub.get("abstract")
    needs_doi = not pub.get("doi")

    if not (needs_title or needs_abstract or needs_doi):
        return pub, False

    cr_data = None

    # Try DOI lookup first (more reliable)
    if pub.get("doi"):
        cr_data = fetch_crossref_by_doi(pub["doi"], email=email)
        time.sleep(CROSSREF_POLITE_DELAY)

    # Fallback: title+year search
    if cr_data is None and pub.get("title") and pub.get("year"):
        cr_data = fetch_crossref_by_title(pub["title"], pub["year"], email=email)
        time.sleep(CROSSREF_POLITE_DELAY)

    if cr_data is None:
        return pub, False

    updated = False

    if needs_title and cr_data.get("title"):
        pub["title"] = cr_data["title"]
        updated = True

    if needs_abstract and cr_data.get("abstract"):
        pub["abstract"] = cr_data["abstract"]
        updated = True

    if needs_doi and cr_data.get("doi"):
        pub["doi"] = cr_data["doi"]
        updated = True

    # Enrich authors if we only had authors_year (parsed names without first names)
    if cr_data.get("authors") and pub.get("authors_year"):
        pub["_crossref_authors"] = cr_data["authors"]
        updated = True

    return pub, updated


def read_publications(path: Path) -> dict[int, dict]:
    """Read publications.xlsx and return {pub_id: {title, doi, abstract, year,
    authors_year}}."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    if ws is None:
        msg = f"No active sheet in {path}"
        raise ValueError(msg)

    # Find column indices from header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {}
    for i, h in enumerate(headers):
        if h is not None:
            col_map[str(h).strip()] = i

    required = {"id", "Authors_year", "year", "title", "doi", "abstract"}
    missing = required - set(col_map.keys())
    if missing:
        msg = f"Missing columns in publications.xlsx: {missing}"
        raise ValueError(msg)

    publications = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pub_id = row[col_map["id"]]
        if pub_id is None:
            continue
        pub_id = int(pub_id)

        title = row[col_map["title"]]
        doi = row[col_map["doi"]]
        abstract = row[col_map["abstract"]]
        year = row[col_map["year"]]
        authors_year = row[col_map["Authors_year"]]

        publications[pub_id] = {
            "title": str(title)[:255] if title else None,
            "doi": str(doi)[:128] if doi and str(doi).strip() else None,
            "abstract": str(abstract)[:8192] if abstract else "",
            "year": str(int(year)) if year else None,
            "authors_year": str(authors_year) if authors_year else "",
        }

    wb.close()
    return publications


def read_sites_publications(path: Path) -> dict[int, list[int]]:
    """Read sites_publications.xlsx and return {pub_id: [site_ids]}."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    if ws is None:
        msg = f"No active sheet in {path}"
        raise ValueError(msg)

    pub_to_sites: dict[int, list[int]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        site_id = row[0]
        pub_id = row[1]
        if site_id is None or pub_id is None:
            continue
        site_id = int(site_id)
        pub_id = int(pub_id)
        pub_to_sites.setdefault(pub_id, []).append(site_id)

    wb.close()
    return pub_to_sites


def read_sites_shapefile(path: Path) -> dict[int, tuple[float, float]]:
    """Read sites.shp and return {site_id: (latitude, longitude)}."""
    gdf = gpd.read_file(path)
    site_coords: dict[int, tuple[float, float]] = {}
    for _, row in gdf.iterrows():
        site_id = int(row["site_id"])
        # Shapefile geometry is Point(lon, lat)
        lon = row.geometry.x
        lat = row.geometry.y
        site_coords[site_id] = (lat, lon)
    return site_coords


def main() -> None:  # noqa: PLR0912, PLR0915
    """Entry point for the import-marcos-data CLI command."""
    parser = argparse.ArgumentParser(
        description="Import Marcos' curated data into the Maress database.",
    )
    parser.add_argument(
        "--email",
        required=True,
        help="Email of the user who will own the imported items.",
    )
    parser.add_argument(
        "--data-dir",
        default="../marcos-custom-data",
        help="Path to marcos-custom-data directory (default: ../marcos-custom-data).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and validate data without writing to the database.",
    )
    parser.add_argument(
        "--no-enrich",
        action="store_true",
        help="Skip CrossRef API lookups for missing metadata.",
    )
    args = parser.parse_args()

    data_dir = Path(args.data_dir)
    pub_path = data_dir / "publications.xlsx"
    sites_pub_path = data_dir / "sites_publications.xlsx"
    shapefile_path = data_dir / "gis" / "sites.shp"

    for p in [pub_path, sites_pub_path, shapefile_path]:
        if not p.exists():
            print(f"ERROR: File not found: {p}")
            sys.exit(1)

    # Read all data sources
    print("Reading publications.xlsx...")
    all_pubs = read_publications(pub_path)
    print(f"  Found {len(all_pubs)} publications")

    print("Reading sites_publications.xlsx...")
    pub_to_sites = read_sites_publications(sites_pub_path)
    print(f"  Found {len(pub_to_sites)} publications with linked sites")

    print("Reading gis/sites.shp...")
    site_coords = read_sites_shapefile(shapefile_path)
    print(f"  Found {len(site_coords)} site coordinates")

    # Filter: every pub that has id AND (doi OR (title AND year))
    eligible_pub_ids = set()
    skipped_ineligible = 0
    for pub_id, pub in all_pubs.items():
        has_doi = bool(pub.get("doi"))
        has_title_year = bool(pub.get("title")) and bool(pub.get("year"))
        if has_doi or has_title_year:
            eligible_pub_ids.add(pub_id)
        else:
            skipped_ineligible += 1

    pub_ids_with_sites = eligible_pub_ids & set(pub_to_sites.keys())
    pub_ids_without_sites = eligible_pub_ids - set(pub_to_sites.keys())

    print(f"\nEligible publications (have DOI or title+year): {len(eligible_pub_ids)}")
    print(f"  With linked sites:    {len(pub_ids_with_sites)}")
    print(f"  Without linked sites: {len(pub_ids_without_sites)}")
    if skipped_ineligible:
        print(f"  Skipped (no DOI and no title+year): {skipped_ineligible}")

    # Check for missing site coordinates
    missing_coords = 0
    for pub_id in pub_ids_with_sites:
        for site_id in pub_to_sites[pub_id]:
            if site_id not in site_coords:
                missing_coords += 1
    if missing_coords:
        print(f"  WARNING: {missing_coords} site references lack coordinates in shapefile")

    # Enrich publications with missing metadata via CrossRef
    enriched_count = 0
    if not args.no_enrich:
        pubs_needing_enrichment = [
            pub_id
            for pub_id in eligible_pub_ids
            if not all_pubs[pub_id].get("title")
            or not all_pubs[pub_id].get("abstract")
            or not all_pubs[pub_id].get("doi")
        ]
        if pubs_needing_enrichment:
            print(f"\nEnriching {len(pubs_needing_enrichment)} publications via CrossRef API...")
            for pub_id in pubs_needing_enrichment:
                pub, was_updated = enrich_publication(all_pubs[pub_id], email=args.email)
                all_pubs[pub_id] = pub
                if was_updated:
                    enriched_count += 1
                    print(f"  Enriched pub_id={pub_id}: {pub.get('title', '?')[:60]}")
            print(f"  Enriched {enriched_count}/{len(pubs_needing_enrichment)} publications")

    if args.dry_run:
        print("\n--- DRY RUN: No database changes will be made ---")
        total_sites = sum(
            len(pub_to_sites.get(pid, []))
            for pid in eligible_pub_ids
        )
        print(f"  Would create {len(eligible_pub_ids)} items")
        print(f"  Would create up to {total_sites} study sites")
        if enriched_count:
            print(f"  Enriched {enriched_count} publications from CrossRef")
        # Show a sample
        sample_id = next(iter(eligible_pub_ids))
        pub = all_pubs[sample_id]
        print(f"\n  Sample publication (id={sample_id}):")
        print(f"    Title: {pub['title']}")
        print(f"    DOI: {pub['doi']}")
        print(f"    Year: {pub['year']}")
        print(f"    Authors: {parse_authors(pub['authors_year'])}")
        print(f"    Sites: {pub_to_sites.get(sample_id, [])}")
        return

    # Database operations
    with SessionLocal() as session:
        # Look up user
        user = crud.get_user_by_email(session=session, email=args.email)
        if not user:
            print(f"ERROR: No user found with email '{args.email}'")
            sys.exit(1)
        print(f"\nImporting as user: {user.full_name} ({user.email})")

        items_created = 0
        items_skipped = 0
        sites_created = 0
        sites_skipped = 0
        errors = 0

        for pub_id in sorted(eligible_pub_ids):
            pub = all_pubs[pub_id]

            # Idempotency: skip if item already exists for this owner
            if pub["doi"]:
                existing = session.exec(
                    select(Item).where(
                        Item.doi == pub["doi"],
                        Item.owner_id == user.id,
                    ),
                ).first()
                if existing:
                    items_skipped += 1
                    print(f"  SKIP (DOI exists): {pub['doi']}")
                    continue
            elif pub["title"] and pub["year"]:
                existing = session.exec(
                    select(Item).where(
                        Item.title == pub["title"],
                        Item.date == pub["year"],
                        Item.owner_id == user.id,
                    ),
                ).first()
                if existing:
                    items_skipped += 1
                    print(f"  SKIP (title+year exists): {pub['title'][:60]}")
                    continue

            try:
                # Create the Item
                key = generate_key()
                item_data = ItemCreate.model_validate(
                    {
                        "key": key,
                        "title": pub["title"],
                        "DOI": pub["doi"],
                        "abstractNote": pub["abstract"],
                        "date": pub["year"],
                        "itemType": "journalArticle",
                    },
                )
                db_item = crud.create_item(
                    session=session,
                    item_in=item_data,
                    owner_id=user.id,
                )
                items_created += 1

                # Create Creators — prefer CrossRef authors (have first names)
                cr_authors = pub.get("_crossref_authors")
                if cr_authors:
                    for author in cr_authors:
                        creator_data = CreatorCreate(
                            lastName=author["lastName"],
                            firstName=author.get("firstName"),
                            creatorType=author["creatorType"],
                        )
                        crud.create_creator(session, creator_data, item_id=db_item.id)
                else:
                    authors = parse_authors(pub["authors_year"])
                    for author in authors:
                        creator_data = CreatorCreate(
                            lastName=author["lastName"],
                            creatorType=author["creatorType"],
                        )
                        crud.create_creator(session, creator_data, item_id=db_item.id)

                # Create StudySites (only for pubs that have linked sites)
                if pub_id in pub_to_sites:
                    for site_id in pub_to_sites[pub_id]:
                        if site_id not in site_coords:
                            sites_skipped += 1
                            continue

                        lat, lon = site_coords[site_id]
                        study_site_data = StudySiteCreate(
                            item_id=db_item.id,
                            latitude=lat,
                            longitude=lon,
                            is_manual=True,
                            confidence_score=1.0,
                            validation_score=1.0,
                            extraction_method=CoordinateExtractionMethod.MANUAL,
                            source_type=CoordinateSourceType.MANUAL,
                            section=PaperSections.OTHER,
                            context="Imported from marcos-custom-data",
                        )
                        crud.create_study_site(session, study_site_data)
                        sites_created += 1

                session.commit()

            except Exception as e:
                session.rollback()
                errors += 1
                print(f"  ERROR (pub_id={pub_id}): {e}")

        print("\n--- Import Summary ---")
        print(f"  Items created:  {items_created}")
        print(f"  Items skipped:  {items_skipped} (DOI already exists)")
        print(f"  Sites created:  {sites_created}")
        print(f"  Sites skipped:  {sites_skipped} (missing coordinates)")
        if enriched_count:
            print(f"  Enriched:       {enriched_count} (via CrossRef)")
        print(f"  Errors:         {errors}")


if __name__ == "__main__":
    main()
